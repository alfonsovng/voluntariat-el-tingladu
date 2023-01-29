from openpyxl import load_workbook
from openpyxl.styles import Alignment
from copy import copy

import os

class ExcelManager:

    def __init__(self, app=None):
        self.excel_absolute_folder_path = None
        self.excel_template_path = None
        if app is not None:
            self.init_app(app)

    def init_app(self, app):
        excel_relative_path = app.config.get('EXCEL_RELATIVE_PATH_FROM_STATIC')
        self.excel_absolute_folder_path = app.static_folder + excel_relative_path
        self.excel_template_path = self.excel_absolute_folder_path + app.config.get("EXCEL_TEMPLATE")

    def create_excel(self, file_name):
        return Excel(
            template_path = self.excel_template_path,
            file_path = self.excel_absolute_folder_path + file_name
        )

    CHUNK_SIZE = 8192
    def stream_and_remove(self, file_name):
        file_path = self.excel_absolute_folder_path + file_name

        with open(file_path, 'rb') as fd:
            while 1:
                buf = fd.read(ExcelManager.CHUNK_SIZE)
                if buf:
                    yield buf
                else:
                    break
            fd.close()
        os.remove(file_path)
    
class Excel:

    MAX_LENGTH = 100

    def __init__(self, template_path, file_path):
        self.file_path = file_path
        self.wb = load_workbook(template_path)
        self.sheet = self.wb["data"]

    def write(self, r: int, values):
        row = r + 1
        for column, value in enumerate(values, start=1):
            cell = self.sheet.cell(row=row, column=column)
            cell.value = value

            if Excel.length(value) > Excel.MAX_LENGTH:
                alignment_obj = copy(cell.alignment)
                alignment_obj.wrap_text=True
                alignment_obj.vertical='top'
                cell.alignment = alignment_obj

                self.sheet.row_dimensions[row].height = 25

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, exc_traceback):
        self.close()

    def close(self):
        # adjust column width
        # https://stackoverflow.com/a/60801712
        for column_cells in self.sheet.columns:
            length = max(Excel.length(cell.value) for cell in column_cells)
            if length > Excel.MAX_LENGTH:
                length = Excel.MAX_LENGTH
            self.sheet.column_dimensions[column_cells[0].column_letter].width = length
    
        self.wb.save(self.file_path)

    def length(text):
        return len(str(text))*1.2